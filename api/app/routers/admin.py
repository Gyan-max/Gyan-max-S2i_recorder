import csv
import hashlib
import io
import json
import os
import random
import zipfile
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Request, status, Query, Response
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete, update, distinct, case
from datetime import datetime

from ..database import get_db
from ..models import Speaker, Clip, Task, Scenario, DeviceSpeaker, WithdrawalAudit
from ..schemas import (
    AdminLoginRequest, AdminLoginResponse, AdminStatsResponse, AdminCoverageResponse,
    AdminCoverageItem, ClipReviewResponse, ClipReviewItem, ClipReviewActionRequest,
    QRGenerateResponse, QRItem, AssignDomainRequest
)
from ..auth import (
    create_admin_token, verify_admin_credentials, get_current_admin,
    check_login_rate_limit, record_failed_login, clear_login_attempts,
)
from ..services.storage import get_export_path

router = APIRouter(prefix="/admin", tags=["Admin"])

@router.post("/login", response_model=AdminLoginResponse)
async def admin_login(req: AdminLoginRequest, request: Request):
    """Authenticates admin credentials and returns a JWT token."""
    client_ip = request.client.host if request.client else "unknown"
    check_login_rate_limit(client_ip)

    if not verify_admin_credentials(req.username, req.password):
        record_failed_login(client_ip)
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )

    clear_login_attempts(client_ip)
    result = create_admin_token(req.username)
    return AdminLoginResponse(token=result["token"], expires_at=result["expires_at"])

@router.get("/stats", response_model=AdminStatsResponse)
async def get_admin_stats(
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Retrieves system-wide statistics for the admin dashboard."""
    # Total speakers (excluding withdrawn)
    stmt_spk = select(func.count(Speaker.speaker_id)).where(Speaker.withdrawn_at == None)
    res_spk = await db.execute(stmt_spk)
    total_speakers = res_spk.scalar() or 0
    
    # Total Clip rows
    stmt_clips = select(func.count(Clip.clip_id))
    res_clips = await db.execute(stmt_clips)
    total_recordings = res_clips.scalar() or 0
    
    # Confirmed / processing / processed clips
    stmt_confirmed = select(func.count(Clip.clip_id)).where(
        Clip.status.in_(["confirmed", "processing", "processed"])
    )
    res_confirmed = await db.execute(stmt_confirmed)
    confirmed_clips = res_confirmed.scalar() or 0
    
    # Redo count (sum of task.redo_count)
    stmt_redo = select(func.sum(Task.redo_count))
    res_redo = await db.execute(stmt_redo)
    redo_count = res_redo.scalar() or 0
    
    # QC passed (status is processed and has no critical qc_flags, or simply status='processed')
    stmt_qc_pass = select(func.count(Clip.clip_id)).where(Clip.status == "processed")
    res_qc_pass = await db.execute(stmt_qc_pass)
    qc_passed = res_qc_pass.scalar() or 0
    
    # QC failed (status is rejected)
    stmt_qc_fail = select(func.count(Clip.clip_id)).where(Clip.status == "rejected")
    res_qc_fail = await db.execute(stmt_qc_fail)
    qc_failed = res_qc_fail.scalar() or 0
    
    return AdminStatsResponse(
        total_speakers=total_speakers,
        total_recordings=total_recordings,
        confirmed_clips=confirmed_clips,
        redo_count=redo_count,
        qc_passed=qc_passed,
        qc_failed=qc_failed
    )

@router.get("/coverage", response_model=AdminCoverageResponse)
async def get_intent_coverage(
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Retrieves the intent coverage heatmap statistics."""
    # Get all scenarios
    scen_stmt = select(Scenario.domain, Scenario.intent)
    scen_res = await db.execute(scen_stmt)
    scenarios = scen_res.all()
    
    # Group into unique (domain, intent) keys
    unique_intents = sorted(list(set((s.domain, s.intent) for s in scenarios)))
    
    coverage_items = []
    for domain, intent in unique_intents:
        # Count processed clips for this intent
        clips_stmt = select(func.count(Clip.clip_id)).join(Task).where(
            Task.intent == intent,
            Clip.status == "processed"
        )
        clips_res = await db.execute(clips_stmt)
        clips_count = clips_res.scalar() or 0
        
        # Count unique speakers who contributed to this intent
        spk_stmt = select(func.count(distinct(Clip.speaker_id))).join(Task).where(
            Task.intent == intent,
            Clip.status == "processed"
        )
        spk_res = await db.execute(spk_stmt)
        spk_count = spk_res.scalar() or 0
        
        coverage_items.append(
            AdminCoverageItem(
                domain=domain,
                intent=intent,
                clips_processed=clips_count,
                speakers_count=spk_count
            )
        )
        
    return AdminCoverageResponse(coverage=coverage_items)

@router.get("/clips", response_model=ClipReviewResponse)
async def get_review_queue_clips(
    status_filter: Optional[str] = Query(None, description="Filter clips by status"),
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Retrieves clips for review, ordering by creation date descending."""
    stmt = select(Clip, Task).join(Task, Clip.task_id == Task.task_id)
    if status_filter:
        stmt = stmt.where(Clip.status == status_filter)
        
    stmt = stmt.order_by(Clip.created_at.desc())
    res = await db.execute(stmt)
    rows = res.all()
    
    items = []
    for clip, task in rows:
        items.append(
            ClipReviewItem(
                clip_id=clip.clip_id,
                task_id=clip.task_id,
                speaker_id=clip.speaker_id,
                device_id=clip.device_id,
                domain=task.domain,
                intent=task.intent,
                scenario_id=task.scenario_id,
                filename=clip.filename,
                duration_s=clip.duration_s,
                qc_flags=clip.qc_flags,
                status=clip.status,
                transcript_provisional=clip.transcript_provisional,
                transcript_final=clip.transcript_final,
                transcript_source=clip.transcript_source,
                created_at=clip.created_at
            )
        )
    return ClipReviewResponse(clips=items)

@router.post("/clips/{clip_id}/review")
async def review_clip_action(
    clip_id: str,
    req: ClipReviewActionRequest,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Allows domain leads to accept, reject, or correct transcripts of clips."""
    stmt = select(Clip).where(Clip.clip_id == clip_id)
    res = await db.execute(stmt)
    clip = res.scalar()
    
    if not clip:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "CLIP_NOT_FOUND", "message": "Clip not found"}
        )
        
    if req.action == "accept":
        clip.status = "processed"
    elif req.action == "reject":
        clip.status = "rejected"
    elif req.action == "edit_transcript":
        if not req.transcript_final:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"code": "TRANSCRIPT_REQUIRED", "message": "transcript_final is required for edit_transcript"}
            )
        clip.transcript_final = req.transcript_final
        clip.transcript_source = "human_verified"
        clip.status = "processed"
        
    await db.commit()
    return {"message": "Clip review completed successfully", "clip_id": clip_id, "status": clip.status}

@router.post("/speakers/{speaker_id}/withdraw")
async def withdraw_speaker(
    speaker_id: str,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Voluntary speaker withdrawal endpoint. Soft-deletes speaker, deletes raw/processed
    audio files, deletes clips and tasks, and writes to withdrawal_audits.
    """
    # 1. Check if speaker exists
    spk_stmt = select(Speaker).where(Speaker.speaker_id == speaker_id)
    spk_res = await db.execute(spk_stmt)
    speaker = spk_res.scalar()
    
    if not speaker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "SPEAKER_NOT_FOUND", "message": "Speaker profile not found"}
        )
        
    if speaker.withdrawn_at is not None:
        return {"message": "Speaker is already withdrawn", "speaker_id": speaker_id}

    # 2. Get counts and file paths for deletion
    clips_stmt = select(Clip).where(Clip.speaker_id == speaker_id)
    clips_res = await db.execute(clips_stmt)
    clips = list(clips_res.scalars().all())
    
    tasks_stmt = select(Task).where(Task.speaker_id == speaker_id)
    tasks_res = await db.execute(tasks_stmt)
    tasks = list(tasks_res.scalars().all())
    
    clips_deleted = len(clips)
    tasks_deleted = len(tasks)
    
    # 3. Physically delete the audio files from storage
    for clip in clips:
        # Delete raw file
        if clip.raw_path and os.path.exists(clip.raw_path):
            try:
                os.remove(clip.raw_path)
            except Exception:
                pass
        # Delete processed WAV file
        if clip.wav_path and os.path.exists(clip.wav_path):
            try:
                os.remove(clip.wav_path)
            except Exception:
                pass
                
    # 4. Perform database deletions and audit logging in a transaction block
    async with db.begin_nested():
        # Soft-delete speaker by setting withdrawn_at
        speaker.withdrawn_at = datetime.utcnow()
        
        # Delete associated records
        await db.execute(delete(Clip).where(Clip.speaker_id == speaker_id))
        await db.execute(delete(Task).where(Task.speaker_id == speaker_id))
        await db.execute(delete(DeviceSpeaker).where(DeviceSpeaker.speaker_id == speaker_id))
        
        # Insert audit log
        audit = WithdrawalAudit(
            speaker_id=speaker_id,
            clips_deleted=clips_deleted,
            tasks_deleted=tasks_deleted,
            processed_by=admin.get("sub"),
            notes="Speaker voluntary withdrawal request"
        )
        db.add(audit)
        
    await db.commit()
    return {
        "message": "Speaker withdrawal completed",
        "speaker_id": speaker_id,
        "clips_deleted": clips_deleted,
        "tasks_deleted": tasks_deleted
    }

@router.post("/speakers/{speaker_id}/assign-domain")
async def assign_speaker_domain(
    speaker_id: str,
    req: AssignDomainRequest,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Assigns a recording domain to a speaker. The speaker will only see tasks from this domain."""
    allowed_domains = {"BNK", "EDU", "TRV", "VAS"}
    if req.domain not in allowed_domains:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "INVALID_DOMAIN", "message": f"Domain must be one of {allowed_domains}"}
        )

    spk_stmt = select(Speaker).where(Speaker.speaker_id == speaker_id)
    spk_res = await db.execute(spk_stmt)
    speaker = spk_res.scalar()

    if not speaker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "SPEAKER_NOT_FOUND", "message": "Speaker not found"}
        )

    speaker.assigned_domain = req.domain
    await db.commit()

    return {"message": f"Domain '{req.domain}' assigned to speaker {speaker_id}", "speaker_id": speaker_id, "assigned_domain": req.domain}

@router.delete("/speakers/{speaker_id}/assign-domain")
async def remove_speaker_domain_assignment(
    speaker_id: str,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Removes the domain assignment from a speaker."""
    spk_stmt = select(Speaker).where(Speaker.speaker_id == speaker_id)
    spk_res = await db.execute(spk_stmt)
    speaker = spk_res.scalar()

    if not speaker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "SPEAKER_NOT_FOUND", "message": "Speaker not found"}
        )

    speaker.assigned_domain = None
    await db.commit()

    return {"message": f"Domain assignment removed for speaker {speaker_id}", "speaker_id": speaker_id}

@router.get("/speakers/{speaker_id}/assignment")
async def get_speaker_domain_assignment(
    speaker_id: str,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Gets the domain assignment for a speaker."""
    spk_stmt = select(Speaker).where(Speaker.speaker_id == speaker_id)
    spk_res = await db.execute(spk_stmt)
    speaker = spk_res.scalar()

    if not speaker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "SPEAKER_NOT_FOUND", "message": "Speaker not found"}
        )

    return {"speaker_id": speaker_id, "assigned_domain": speaker.assigned_domain}

@router.get("/export")
async def export_dataset(
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Generates a speaker-disjoint dataset split (80/10/10 ratio) and returns a CSV manifest.
    Ensures that no speaker's recordings appear in multiple splits.
    """
    # 1. Fetch all processed clips with metadata
    stmt = (
        select(Clip, Speaker, Task)
        .join(Speaker, Clip.speaker_id == Speaker.speaker_id)
        .join(Task, Clip.task_id == Task.task_id)
        .where(
            Clip.status == "processed",
            Speaker.withdrawn_at == None
        )
    )
    res = await db.execute(stmt)
    rows = res.all()
    
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "NO_DATA", "message": "No processed clips available for export"}
        )
        
    # 2. Assign speaker-disjoint splits.
    # sorted() before shuffling is required for reproducibility: set iteration
    # order varies between processes because Python salts string hashing, so
    # seeding the RNG alone is not enough.
    speakers = sorted({r.Speaker.speaker_id for r in rows})
    random.Random(42).shuffle(speakers)

    n_speakers = len(speakers)

    # Proportional splits break down at small speaker counts, so guarantee a
    # non-empty train set and only carve out dev/test when there are enough
    # speakers to keep them disjoint.
    if n_speakers == 1:
        train_count, dev_count = 1, 0
    elif n_speakers == 2:
        train_count, dev_count = 1, 0  # remaining speaker becomes test
    elif n_speakers < 10:
        # One speaker each for dev and test, the rest train.
        train_count, dev_count = n_speakers - 2, 1
    else:
        train_count = max(1, round(n_speakers * 0.8))
        dev_count = max(1, round(n_speakers * 0.1))
        # Never let rounding consume every speaker; test needs at least one.
        if train_count + dev_count >= n_speakers:
            dev_count = max(1, n_speakers - train_count - 1)
            train_count = n_speakers - dev_count - 1

    speaker_splits = {}
    for idx, spk in enumerate(speakers):
        if idx < train_count:
            speaker_splits[spk] = "train"
        elif idx < train_count + dev_count:
            speaker_splits[spk] = "dev"
        else:
            speaker_splits[spk] = "test"

    # Each speaker maps to exactly one split by construction.
    
    # 3. Create in-memory CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # CSV Header
    writer.writerow([
        "clip_id", "filename", "speaker_id", "gender", "age_band", 
        "l1", "region", "domain", "intent", "scenario_id", 
        "scenario_no", "example_no", "prompted", "duration_s", 
        "transcript_final", "transcript_source", "qc_flags", "split"
    ])
    
    for clip, speaker, task in rows:
        split = speaker_splits.get(speaker.speaker_id, "train")
        qc_flags_str = ",".join(clip.qc_flags) if clip.qc_flags else ""
        writer.writerow([
            clip.clip_id,
            clip.filename or "",
            speaker.speaker_id,
            speaker.gender,
            speaker.age_band,
            speaker.l1,
            speaker.region,
            task.domain,
            task.intent,
            task.scenario_id,
            task.scenario_no,
            task.example_no,
            clip.prompted,
            clip.duration_s or 0.0,
            clip.transcript_final or "",
            clip.transcript_source or "",
            qc_flags_str,
            split
        ])
        
    csv_content = output.getvalue()
    output.close()
    
    # Save the manifest.csv locally in storage/exports
    os.makedirs(os.path.dirname(get_export_path("manifest.csv")), exist_ok=True)
    with open(get_export_path("manifest.csv"), "w") as f:
        f.write(csv_content)
        
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=manifest.csv"}
    )

@router.get("/export/excel")
async def export_dataset_excel(
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Exports dataset in Excel format with multiple sheets for better organization.
    Includes: Speakers, Recordings, Coverage, and QC Summary sheets.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"code": "DEPENDENCY_ERROR", "message": "openpyxl not installed. Run: pip install openpyxl"}
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Style definitions
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Sheet 1: Speakers Data
    ws_speakers = wb.create_sheet("Speakers")
    speaker_headers = ["Speaker ID", "Gender", "Age Band", "L1", "Region", "Total Clips", "Confirmed Clips", "Consent Date"]
    ws_speakers.append(speaker_headers)
    
    # Style header
    for col_num, _ in enumerate(speaker_headers, 1):
        cell = ws_speakers.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Fetch speaker data
    speakers_stmt = select(
        Speaker.speaker_id,
        Speaker.gender,
        Speaker.age_band,
        Speaker.l1,
        Speaker.region,
        Speaker.consent_at,
        func.count(Clip.clip_id).label("total_clips"),
        func.sum(case((Clip.status == "confirmed", 1), else_=0)).label("confirmed_clips")
    ).outerjoin(Clip).where(Speaker.withdrawn_at == None).group_by(Speaker.speaker_id)
    
    speakers_res = await db.execute(speakers_stmt)
    for row in speakers_res:
        ws_speakers.append([
            row.speaker_id,
            row.gender,
            row.age_band,
            row.l1,
            row.region,
            row.total_clips or 0,
            row.confirmed_clips or 0,
            row.consent_at.strftime("%Y-%m-%d %H:%M") if row.consent_at else ""
        ])
    
    # Auto-adjust column widths
    for col in ws_speakers.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_speakers.column_dimensions[column].width = adjusted_width
    
    # Sheet 2: All Recordings
    ws_clips = wb.create_sheet("Recordings")
    clip_headers = ["Clip ID", "Speaker ID", "Domain", "Intent", "Scenario", "Status", "Duration (s)", 
                    "QC Flags", "Transcript", "Prompted", "Created At"]
    ws_clips.append(clip_headers)
    
    for col_num, _ in enumerate(clip_headers, 1):
        cell = ws_clips.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    clips_stmt = (
        select(Clip, Task)
        .join(Task, Clip.task_id == Task.task_id)
        .order_by(Clip.created_at.desc())
    )
    clips_res = await db.execute(clips_stmt)
    
    for clip, task in clips_res:
        qc_flags_str = ", ".join(clip.qc_flags) if clip.qc_flags else "None"
        ws_clips.append([
            clip.clip_id,
            clip.speaker_id,
            task.domain,
            task.intent,
            task.scenario_id,
            clip.status,
            clip.duration_s or 0.0,
            qc_flags_str,
            clip.transcript_final or clip.transcript_provisional or "",
            "Yes" if clip.prompted else "No",
            clip.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    for col in ws_clips.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_clips.column_dimensions[column].width = adjusted_width
    
    # Sheet 3: Coverage by Intent
    ws_coverage = wb.create_sheet("Coverage")
    coverage_headers = ["Domain", "Intent", "Clips Processed", "Unique Speakers", "Target", "Progress %"]
    ws_coverage.append(coverage_headers)
    
    for col_num, _ in enumerate(coverage_headers, 1):
        cell = ws_coverage.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get coverage data
    scenarios_stmt = select(Scenario.domain, Scenario.intent).distinct()
    scenarios_res = await db.execute(scenarios_stmt)
    unique_intents = sorted(list(set((s.domain, s.intent) for s in scenarios_res)))
    
    for domain, intent in unique_intents:
        clips_count_stmt = select(func.count(Clip.clip_id)).join(Task).where(
            Task.intent == intent,
            Clip.status == "processed"
        )
        clips_count_res = await db.execute(clips_count_stmt)
        clips_count = clips_count_res.scalar() or 0
        
        speakers_count_stmt = select(func.count(distinct(Clip.speaker_id))).join(Task).where(
            Task.intent == intent,
            Clip.status == "processed"
        )
        speakers_count_res = await db.execute(speakers_count_stmt)
        speakers_count = speakers_count_res.scalar() or 0
        
        target = 40
        progress = min(100, (clips_count / target * 100)) if target > 0 else 0
        
        ws_coverage.append([
            domain,
            intent,
            clips_count,
            speakers_count,
            target,
            f"{progress:.1f}%"
        ])
    
    for col in ws_coverage.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_coverage.column_dimensions[column].width = adjusted_width
    
    # Sheet 4: QC Summary
    ws_qc = wb.create_sheet("QC Summary")
    qc_headers = ["QC Flag", "Count", "Percentage"]
    ws_qc.append(qc_headers)
    
    for col_num, _ in enumerate(qc_headers, 1):
        cell = ws_qc.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get total clips
    total_clips_stmt = select(func.count(Clip.clip_id))
    total_clips_res = await db.execute(total_clips_stmt)
    total_clips = total_clips_res.scalar() or 0
    
    # Count clips by QC flags (this is simplified - in production you'd parse the array)
    qc_summary = {
        "No Issues": 0,
        "Has QC Flags": 0,
        "Rejected": 0
    }
    
    clips_qc_stmt = select(Clip.status, Clip.qc_flags)
    clips_qc_res = await db.execute(clips_qc_stmt)
    
    for clip_row in clips_qc_res:
        if clip_row.status == "rejected":
            qc_summary["Rejected"] += 1
        elif clip_row.qc_flags and len(clip_row.qc_flags) > 0:
            qc_summary["Has QC Flags"] += 1
        else:
            qc_summary["No Issues"] += 1
    
    for flag, count in qc_summary.items():
        percentage = (count / total_clips * 100) if total_clips > 0 else 0
        ws_qc.append([flag, count, f"{percentage:.1f}%"])
    
    for col in ws_qc.columns:
        ws_qc.column_dimensions[col[0].column_letter].width = 20
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Also save locally
    export_dir = os.path.dirname(get_export_path("dataset_export.xlsx"))
    os.makedirs(export_dir, exist_ok=True)
    with open(get_export_path("dataset_export.xlsx"), "wb") as f:
        f.write(excel_file.getvalue())
    excel_file.seek(0)
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dataset_export.xlsx"}
    )


@router.get("/export/research-bundle")
async def export_research_bundle(
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Download a self-contained, research-ready ZIP archive.

    The archive contains processed WAV recordings, an Excel workbook, a CSV
    manifest, task/prompt metadata, checksums, and a README. Only processed
    clips belonging to non-withdrawn speakers are included.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"code": "DEPENDENCY_ERROR", "message": "openpyxl is required for research exports"}
        )

    stmt = (
        select(Clip, Speaker, Task, Scenario)
        .join(Speaker, Clip.speaker_id == Speaker.speaker_id)
        .join(Task, Clip.task_id == Task.task_id)
        .join(Scenario, Task.scenario_id == Scenario.scenario_id)
        .where(Clip.status == "processed", Speaker.withdrawn_at == None)
        .order_by(Clip.created_at)
    )
    rows = (await db.execute(stmt)).all()
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "NO_DATA", "message": "No processed clips are available for research export"}
        )

    # A stable speaker-disjoint split protects against speaker leakage between sets.
    speaker_ids = sorted({speaker.speaker_id for _, speaker, _, _ in rows})
    speaker_splits = {}
    for index, speaker_id in enumerate(speaker_ids):
        bucket = int(hashlib.sha256(speaker_id.encode()).hexdigest(), 16) % 100
        speaker_splits[speaker_id] = "train" if bucket < 80 else "dev" if bucket < 90 else "test"

    exported_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    manifest_columns = [
        "clip_id", "audio_path", "sha256", "speaker_id", "split", "gender", "age_band",
        "l1", "region", "domain", "intent", "scenario_id", "scenario_no", "example_no",
        "prompt_text", "register", "prompted", "duration_s", "transcript", "transcript_source",
        "qc_flags", "created_at"
    ]
    manifest_rows = []
    task_rows = {}
    speaker_rows = {}
    missing_audio = []
    audio_files = []

    for clip, speaker, task, scenario in rows:
        audio_path = clip.wav_path if clip.wav_path and os.path.exists(clip.wav_path) else None
        if not audio_path:
            missing_audio.append({"clip_id": clip.clip_id, "expected_path": clip.wav_path, "reason": "processed_audio_missing"})
            continue

        filename = os.path.basename(audio_path)
        archive_path = f"audio/{speaker_splits[speaker.speaker_id]}/{filename}"
        with open(audio_path, "rb") as audio_file:
            checksum = hashlib.sha256(audio_file.read()).hexdigest()

        manifest_rows.append({
            "clip_id": clip.clip_id,
            "audio_path": archive_path,
            "sha256": checksum,
            "speaker_id": speaker.speaker_id,
            "split": speaker_splits[speaker.speaker_id],
            "gender": speaker.gender,
            "age_band": speaker.age_band,
            "l1": speaker.l1,
            "region": speaker.region,
            "domain": task.domain,
            "intent": task.intent,
            "scenario_id": task.scenario_id,
            "scenario_no": task.scenario_no,
            "example_no": task.example_no,
            "prompt_text": scenario.text_hi,
            "register": scenario.register or "",
            "prompted": clip.prompted,
            "duration_s": clip.duration_s or "",
            "transcript": clip.transcript_final or clip.transcript_provisional or "",
            "transcript_source": clip.transcript_source or "",
            "qc_flags": ",".join(clip.qc_flags or []),
            "created_at": clip.created_at.isoformat() if clip.created_at else ""
        })
        audio_files.append((audio_path, archive_path))
        task_rows[task.task_id] = {
            "task_id": task.task_id, "domain": task.domain, "intent": task.intent,
            "scenario_id": task.scenario_id, "scenario_no": task.scenario_no,
            "example_no": task.example_no, "prompt_text": scenario.text_hi,
            "examples": " | ".join(scenario.examples or []), "register": scenario.register or ""
        }
        speaker_rows[speaker.speaker_id] = {
            "speaker_id": speaker.speaker_id, "split": speaker_splits[speaker.speaker_id],
            "gender": speaker.gender, "age_band": speaker.age_band, "l1": speaker.l1,
            "region": speaker.region, "consent_version": speaker.consent_version or "",
            "consent_at": speaker.consent_at.isoformat() if speaker.consent_at else ""
        }

    if not manifest_rows:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "NO_AUDIO", "message": "Processed clip records exist, but their audio files are unavailable"}
        )

    def csv_bytes(columns, records):
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
        return output.getvalue().encode("utf-8")

    # The workbook makes speakers, recordings, prompts, and QC data directly usable in Excel.
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    workbook_sheets = {
        "Recordings": (manifest_columns, manifest_rows),
        "Speakers": (list(next(iter(speaker_rows.values())).keys()), list(speaker_rows.values())),
        "Tasks": (list(next(iter(task_rows.values())).keys()), list(task_rows.values())),
        "Missing audio": (["clip_id", "expected_path", "reason"], missing_audio),
    }
    for sheet_name, (columns, records) in workbook_sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(columns)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for record in records:
            sheet.append([record.get(column, "") for column in columns])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(width + 2, 48)
    excel_output = io.BytesIO()
    workbook.save(excel_output)

    metadata = {
        "format_version": "1.0",
        "exported_at": exported_at,
        "exported_by": admin.get("sub", "admin") if isinstance(admin, dict) else str(admin),
        "selection": "processed clips only; withdrawn speakers excluded",
        "audio_format": "16 kHz, mono PCM WAV",
        "recordings_included": len(manifest_rows),
        "speakers_included": len(speaker_rows),
        "missing_audio_files": len(missing_audio),
        "split_method": "speaker-disjoint stable SHA-256 bucket (80/10/10 target)"
    }
    readme = """# Hinglish S2I research export\n\nThis archive is self-contained for local research storage.\n\n- `audio/` contains processed 16 kHz mono WAV recordings, organised by speaker-disjoint split.\n- `metadata/research_dataset.xlsx` contains Recordings, Speakers, Tasks and Missing audio sheets.\n- `metadata/manifest.csv` is the machine-readable recording manifest.\n- `metadata/export_metadata.json` documents this export.\n\nUse `sha256` in the manifest to verify every local audio file. Do not attempt to re-identify speakers or distribute this archive outside the approved research workflow.\n"""

    archive = io.BytesIO()
    with zipfile.ZipFile(archive, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        for source_path, archive_path in audio_files:
            zip_file.write(source_path, archive_path)
        zip_file.writestr("metadata/manifest.csv", csv_bytes(manifest_columns, manifest_rows))
        zip_file.writestr("metadata/speakers.csv", csv_bytes(list(next(iter(speaker_rows.values())).keys()), list(speaker_rows.values())))
        zip_file.writestr("metadata/tasks.csv", csv_bytes(list(next(iter(task_rows.values())).keys()), list(task_rows.values())))
        zip_file.writestr("metadata/research_dataset.xlsx", excel_output.getvalue())
        zip_file.writestr("metadata/export_metadata.json", json.dumps(metadata, indent=2))
        zip_file.writestr("README.md", readme)

    filename = f"hinglish_s2i_research_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"
    return Response(
        content=archive.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/speakers/detailed")
async def get_speakers_detailed(
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Returns detailed speaker information including recording statistics and demographics.
    """
    stmt = select(
        Speaker,
        func.count(Clip.clip_id).label("total_clips"),
        func.sum(case((Clip.status == "confirmed", 1), else_=0)).label("confirmed_clips"),
        func.sum(case((Clip.status == "processed", 1), else_=0)).label("processed_clips"),
        func.sum(case((Clip.status == "rejected", 1), else_=0)).label("rejected_clips"),
        func.avg(Clip.duration_s).label("avg_duration")
    ).outerjoin(Clip).where(Speaker.withdrawn_at == None).group_by(Speaker.speaker_id).order_by(Speaker.created_at.desc())
    
    res = await db.execute(stmt)
    rows = res.all()
    
    speakers_list = []
    for row in rows:
        speakers_list.append({
            "speaker_id": row.Speaker.speaker_id,
            "name": row.Speaker.name,
            "gender": row.Speaker.gender,
            "age": row.Speaker.age,
            "age_band": row.Speaker.age_band,
            "l1": row.Speaker.l1,
            "region": row.Speaker.region,
            "consent_at": row.Speaker.consent_at.isoformat() if row.Speaker.consent_at else None,
            "created_at": row.Speaker.created_at.isoformat(),
            "total_clips": row.total_clips or 0,
            "confirmed_clips": row.confirmed_clips or 0,
            "processed_clips": row.processed_clips or 0,
            "rejected_clips": row.rejected_clips or 0,
            "avg_duration": round(row.avg_duration, 2) if row.avg_duration else 0.0,
            "assigned_domain": row.Speaker.assigned_domain
        })
    
    return {"speakers": speakers_list}

@router.get("/clips/{clip_id}/audio")
async def get_clip_audio(
    clip_id: str,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Serves the audio file for a specific clip for admin playback.
    """
    stmt = select(Clip).where(Clip.clip_id == clip_id)
    res = await db.execute(stmt)
    clip = res.scalar()
    
    if not clip:
        raise HTTPException(status_code=404, detail="Clip not found")
    
    # Try to serve processed WAV first, then fall back to raw
    audio_path = None
    if clip.wav_path and os.path.exists(clip.wav_path):
        audio_path = clip.wav_path
    elif clip.raw_path and os.path.exists(clip.raw_path):
        audio_path = clip.raw_path
    
    if not audio_path:
        raise HTTPException(status_code=404, detail="Audio file not found")
    
    return FileResponse(audio_path, media_type="audio/wav" if audio_path.endswith(".wav") else "audio/webm")
