"""
Admin endpoints.

Guarded by the `admin: true` custom claim rather than a shared password, so
access is per-person and revocable without a redeploy.
"""

import csv
import io
import logging
import zipfile
from collections import defaultdict
from datetime import datetime, timezone

from firebase_admin import auth as fb_auth
from flask import Blueprint, jsonify, request

from .. import config
from ..auth import admin_required
from ..db import db, delete_object, download_bytes, get_doc, query_all
from ..errors import ApiError
from ..services.clip_deletion import ACTOR_ADMIN, delete_clip_completely

logger = logging.getLogger(__name__)
bp = Blueprint("admin", __name__)


def _now():
    return datetime.now(timezone.utc)


def _iso(value):
    return value.isoformat() if hasattr(value, "isoformat") else value


# ==================== Stats & coverage ====================

@bp.get("/admin/stats")
@admin_required
def stats():
    clips = query_all(config.CLIPS)
    speakers = [s for s in query_all(config.SPEAKERS) if not s.get("withdrawn_at")]
    tasks = query_all(config.TASKS)

    return jsonify({
        "total_speakers": len(speakers),
        "total_recordings": len(clips),
        "confirmed_clips": sum(
            1 for c in clips if c.get("status") in ("confirmed", "processing", "processed")
        ),
        "redo_count": sum(int(t.get("redo_count", 0) or 0) for t in tasks),
        "qc_passed": sum(1 for c in clips if c.get("status") == "processed"),
        "qc_failed": sum(1 for c in clips if c.get("status") == "rejected"),
    })


@bp.get("/admin/coverage")
@admin_required
def coverage():
    """Per-intent progress toward the collection floor."""
    clips = query_all(config.CLIPS)
    usable = [c for c in clips if c.get("status") in ("confirmed", "processing", "processed")]

    grouped = defaultdict(list)
    for c in usable:
        grouped[(c.get("domain"), c.get("intent"))].append(c)

    # Include intents with zero clips so gaps are visible, not just absent.
    for s in query_all(config.SCENARIOS):
        grouped.setdefault((s.get("domain"), s.get("intent")), [])

    items = []
    for (domain, intent), group in sorted(grouped.items(), key=lambda kv: (kv[0][0] or "", kv[0][1] or "")):
        items.append({
            "domain": domain,
            "intent": intent,
            "clips_processed": len(group),
            "speakers_count": len({c.get("speaker_id") for c in group}),
            "floor": config.COVERAGE_FLOOR,
        })
    return jsonify({"coverage": items})


# ==================== Clip review ====================

@bp.get("/admin/clips")
@admin_required
def list_clips():
    status_filter = request.args.get("status_filter")
    filters = [("status", "==", status_filter)] if status_filter else []
    clips = query_all(config.CLIPS, filters)
    clips.sort(key=lambda c: c.get("created_at") or 0, reverse=True)

    return jsonify({"clips": [
        {
            "clip_id": c["clip_id"],
            "task_id": c.get("task_id"),
            "speaker_id": c.get("speaker_id"),
            "device_id": c.get("device_id") or "",
            "domain": c.get("domain"),
            "intent": c.get("intent"),
            "scenario_id": c.get("scenario_id"),
            "filename": c.get("filename"),
            "duration_s": c.get("duration_s"),
            "qc_flags": c.get("qc_flags") or [],
            "status": c.get("status"),
            "transcript_provisional": c.get("transcript_provisional"),
            "transcript_final": c.get("transcript_final"),
            "transcript_source": c.get("transcript_source"),
            "created_at": _iso(c.get("created_at")),
        }
        for c in clips
    ]})


@bp.post("/admin/clips/<clip_id>/review")
@admin_required
def review_clip(clip_id):
    """Accept, reject, or correct a transcript."""
    body = request.get_json(silent=True) or {}
    action = body.get("action")

    clip = get_doc(config.CLIPS, clip_id)
    if not clip:
        raise ApiError(404, "CLIP_NOT_FOUND", "Clip not found.")

    ref = db().collection(config.CLIPS).document(clip_id)

    if action == "accept":
        ref.update({"status": "processed"})
        return jsonify({"clip_id": clip_id, "status": "processed"})

    if action == "reject":
        ref.update({"status": "rejected"})
        return jsonify({"clip_id": clip_id, "status": "rejected"})

    if action == "edit_transcript":
        text = (body.get("transcript_final") or "").strip()
        if not text:
            raise ApiError(400, "EMPTY_TRANSCRIPT", "Transcript cannot be empty.")
        # human_verified outranks ASR and prompt-derived text downstream.
        ref.update({"transcript_final": text, "transcript_source": "human_verified"})
        return jsonify({"clip_id": clip_id, "status": clip.get("status")})

    raise ApiError(400, "UNKNOWN_ACTION", f"Unsupported review action: {action}")


@bp.get("/admin/clips/<clip_id>/audio")
@admin_required
def clip_audio(clip_id):
    """Streams clip audio for review - processed WAV if ready, else raw."""
    clip = get_doc(config.CLIPS, clip_id)
    if not clip:
        raise ApiError(404, "CLIP_NOT_FOUND", "Clip not found.")

    path = clip.get("wav_path") or clip.get("raw_path")
    data = download_bytes(path) if path else None
    if data is None:
        raise ApiError(404, "AUDIO_NOT_FOUND", "Audio file not found.")

    return data, 200, {"Content-Type": "audio/wav" if str(path).endswith(".wav") else "audio/webm"}


@bp.delete("/admin/clips/<clip_id>")
@admin_required
def delete_clip(clip_id):
    """
    Permanently deletes one recording, whoever it belongs to.

    Unlike withdrawal this touches a single clip and leaves the speaker's
    profile and other recordings intact.
    """
    clip = get_doc(config.CLIPS, clip_id)
    if not clip:
        return jsonify({"clip_id": clip_id, "deleted": True})
    delete_clip_completely(clip, actor=ACTOR_ADMIN)
    return jsonify({"clip_id": clip_id, "deleted": True})


# ==================== Speakers ====================

@bp.get("/admin/speakers/detailed")
@admin_required
def speakers_detailed():
    speakers = query_all(config.SPEAKERS)
    clips = query_all(config.CLIPS)

    by_speaker = defaultdict(list)
    for c in clips:
        by_speaker[c.get("speaker_id")].append(c)

    items = []
    for s in sorted(speakers, key=lambda x: x.get("speaker_id", "")):
        own = by_speaker.get(s.get("speaker_id"), [])
        durations = [c["duration_s"] for c in own if c.get("duration_s")]
        items.append({
            "speaker_id": s.get("speaker_id"),
            "name": s.get("name"),
            "gender": s.get("gender"),
            "age": s.get("age"),
            "age_band": s.get("age_band"),
            "l1": s.get("l1"),
            "region": s.get("region"),
            "consent_at": _iso(s.get("consent_at")),
            "created_at": _iso(s.get("created_at")),
            "assigned_domain": s.get("assigned_domain"),
            "withdrawn_at": _iso(s.get("withdrawn_at")),
            "total_clips": len(own),
            "confirmed_clips": sum(1 for c in own if c.get("status") == "confirmed"),
            "processed_clips": sum(1 for c in own if c.get("status") == "processed"),
            "rejected_clips": sum(1 for c in own if c.get("status") == "rejected"),
            "avg_duration": (sum(durations) / len(durations)) if durations else 0.0,
        })
    return jsonify({"speakers": items})


@bp.post("/admin/speakers/<speaker_id>/withdraw")
@admin_required
def withdraw_speaker(speaker_id):
    """
    Right-to-erasure: deletes every recording and task for a speaker and
    disables their sign-in, retaining only an anonymised audit record.
    """
    matches = query_all(config.SPEAKERS, [("speaker_id", "==", speaker_id)])
    if not matches:
        raise ApiError(404, "SPEAKER_NOT_FOUND", "Speaker not found.")
    speaker = matches[0]

    clips = query_all(config.CLIPS, [("speaker_id", "==", speaker_id)])
    tasks = query_all(config.TASKS, [("speaker_id", "==", speaker_id)])

    for c in clips:
        delete_object(c.get("raw_path"))
        delete_object(c.get("wav_path"))

    client = db()
    writer = client.batch()
    for c in clips:
        writer.delete(client.collection(config.CLIPS).document(c["clip_id"]))
    for t in tasks:
        writer.delete(client.collection(config.TASKS).document(t["task_id"]))
    writer.commit()

    client.collection(config.SPEAKERS).document(speaker["id"]).update({
        "withdrawn_at": _now(),
        # Demographics go too - the audit record is deliberately anonymous.
        "name": None, "email": None, "age": None, "gender": None,
        "l1": None, "region": None,
    })

    # Revoke the account so an existing ID token cannot keep working.
    uid = speaker.get("uid") or speaker.get("id")
    try:
        fb_auth.update_user(uid, disabled=True)
        fb_auth.revoke_refresh_tokens(uid)
    except Exception as e:
        logger.warning("Could not disable auth user %s: %s", uid, e)

    client.collection(config.WITHDRAWAL_AUDITS).document().set({
        "speaker_id": speaker_id,
        "clips_deleted": len(clips),
        "tasks_deleted": len(tasks),
        "withdrawn_at": _now(),
    })

    return jsonify({
        "speaker_id": speaker_id,
        "clips_deleted": len(clips),
        "tasks_deleted": len(tasks),
    })


@bp.post("/admin/speakers/<speaker_id>/assign-domain")
@admin_required
def assign_domain(speaker_id):
    body = request.get_json(silent=True) or {}
    domain = body.get("domain")
    if domain not in config.DOMAINS:
        raise ApiError(400, "INVALID_DOMAIN", f"Domain must be one of {config.DOMAINS}.")

    matches = query_all(config.SPEAKERS, [("speaker_id", "==", speaker_id)])
    if not matches:
        raise ApiError(404, "SPEAKER_NOT_FOUND", "Speaker not found.")

    db().collection(config.SPEAKERS).document(matches[0]["id"]).update({"assigned_domain": domain})
    return jsonify({"speaker_id": speaker_id, "assigned_domain": domain})


@bp.delete("/admin/speakers/<speaker_id>/assign-domain")
@admin_required
def clear_assigned_domain(speaker_id):
    matches = query_all(config.SPEAKERS, [("speaker_id", "==", speaker_id)])
    if not matches:
        raise ApiError(404, "SPEAKER_NOT_FOUND", "Speaker not found.")
    db().collection(config.SPEAKERS).document(matches[0]["id"]).update({"assigned_domain": None})
    return jsonify({"speaker_id": speaker_id, "assigned_domain": None})


@bp.get("/admin/speakers/<speaker_id>/assignment")
@admin_required
def get_assignment(speaker_id):
    matches = query_all(config.SPEAKERS, [("speaker_id", "==", speaker_id)])
    if not matches:
        raise ApiError(404, "SPEAKER_NOT_FOUND", "Speaker not found.")
    return jsonify({
        "speaker_id": speaker_id,
        "assigned_domain": matches[0].get("assigned_domain"),
    })


@bp.post("/admin/grant")
@admin_required
def grant_admin():
    """
    Promotes another Firebase user to admin.

    Bootstrapping the first admin cannot happen here (it would be an open
    door); use scripts/set_admin.py with service-account credentials.
    """
    body = request.get_json(silent=True) or {}
    email = (body.get("email") or "").strip()
    if not email:
        raise ApiError(400, "EMAIL_REQUIRED", "An email address is required.")
    try:
        user = fb_auth.get_user_by_email(email)
    except Exception:
        raise ApiError(404, "USER_NOT_FOUND", f"No Firebase user with email {email}.")

    fb_auth.set_custom_user_claims(user.uid, {"admin": True})
    # Force a token refresh so the claim takes effect on their next request.
    fb_auth.revoke_refresh_tokens(user.uid)
    return jsonify({"email": email, "uid": user.uid, "admin": True})


# ==================== Export ====================

def _manifest_rows():
    """One row per corpus-ready clip. age_band only - raw age never exports."""
    clips = [c for c in query_all(config.CLIPS) if c.get("status") == "processed"]
    speakers = {s.get("speaker_id"): s for s in query_all(config.SPEAKERS)}

    rows = []
    for c in clips:
        s = speakers.get(c.get("speaker_id")) or {}
        if s.get("withdrawn_at"):
            continue
        rows.append({
            "filename": c.get("filename"),
            "wav_path": c.get("wav_path"),
            "domain": c.get("domain"),
            "intent": c.get("intent"),
            "scenario_id": c.get("scenario_id"),
            "speaker_id": c.get("speaker_id"),
            "age_band": s.get("age_band"),
            "gender": s.get("gender"),
            "l1": s.get("l1"),
            "region": s.get("region"),
            "transcript": c.get("transcript_final") or c.get("transcript_provisional"),
            "transcript_source": c.get("transcript_source"),
            "prompted": c.get("prompted"),
            "duration_s": c.get("duration_s"),
            "qc_flags": ";".join(c.get("qc_flags") or []),
        })
    return rows


@bp.get("/admin/export")
@admin_required
def export_csv():
    rows = _manifest_rows()
    buf = io.StringIO()
    fields = list(rows[0].keys()) if rows else ["filename"]
    writer = csv.DictWriter(buf, fieldnames=fields)
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue(), 200, {
        "Content-Type": "text/csv",
        "Content-Disposition": 'attachment; filename="manifest.csv"',
    }


@bp.get("/admin/export/excel")
@admin_required
def export_excel():
    """Formatted workbook of the same manifest rows the CSV export produces."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = _manifest_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Clips"

    fields = list(rows[0].keys()) if rows else ["filename"]
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(f) for f in fields])

    # Rough auto-width so the sheet is readable without manual resizing.
    for idx, field in enumerate(fields, 1):
        longest = max([len(str(field))] + [len(str(r.get(field) or "")) for r in rows] or [10])
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(longest + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), 200, {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": 'attachment; filename="dataset_export.xlsx"',
    }


@bp.get("/admin/export/research-bundle")
@admin_required
def export_bundle():
    """Processed WAVs plus the manifest, zipped in memory."""
    rows = _manifest_rows()
    buf = io.BytesIO()

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        manifest = io.StringIO()
        if rows:
            writer = csv.DictWriter(manifest, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        zf.writestr("manifest.csv", manifest.getvalue())

        missing = []
        for row in rows:
            path = row.get("wav_path")
            data = download_bytes(path) if path else None
            if data is None:
                missing.append(row.get("filename"))
                continue
            zf.writestr(f"audio/{row['filename']}", data)

        zf.writestr(
            "README.txt",
            "Hinglish S2I corpus export\n"
            f"Generated: {_now().isoformat()}\n"
            f"Clips: {len(rows)}\n"
            f"Missing audio: {len(missing)}\n\n"
            "Audio is 16kHz mono WAV. age_band is included; raw age is never exported.\n",
        )

    return buf.getvalue(), 200, {
        "Content-Type": "application/zip",
        "Content-Disposition": 'attachment; filename="hinglish_s2i_research_export.zip"',
    }
